#!/usr/bin/env python3
"""xlsx_import.py — read a cross-check workbook back into a results.js payload.

The lab upload path uses this: parse Trades_BS -> recompute every metric with the
CANONICAL engine (engine.metrics, Rule 6B) -> compare against the Claude_Summary
sheet -> emit (results_payload, crosscheck). The lab then renders results_payload
through the SAME dashboard the hunt runs use, and shows the crosscheck banner.

The recompute is the independent second eye: if the trades don't sum to the claimed
net, or "net %" is being read as a CAGR (TRAP-127), the delta surfaces here.

Pure read/compute — no order/risk/live path.
"""
import os
import sys
import openpyxl
import pandas as pd
import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import engine

# tolerance for "match" per metric (abs)
_TOL = {"sharpe": 0.02, "profit_factor": 0.02, "win_rate": 0.2, "maxdd": 0.2,
        "net_pct": 0.2, "net_abs": 5.0, "expectancy": 5.0}


def jsonable(o):
    """Recursively make a payload JSON/JS-safe: numpy scalars -> py, NaN/Inf -> None."""
    import math
    if isinstance(o, dict):
        return {str(k): jsonable(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        return [jsonable(v) for v in o]
    if isinstance(o, float):
        return None if (math.isnan(o) or math.isinf(o)) else o
    if hasattr(o, "item"):                      # numpy scalar
        try:
            return jsonable(o.item())
        except Exception:
            return None
    return o


def _rows(ws):
    head = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out.append(dict(zip(head, r)))
    return out


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse(path):
    """Returns dict: {meta, trades, claude_summary}. trades = all_trades-shaped."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    names = wb.sheetnames
    if "Trades_BS" not in names:
        raise ValueError("workbook has no 'Trades_BS' sheet")

    meta = {}
    if "Meta" in names:
        for r in _rows(wb["Meta"]):
            meta[str(r.get("Field"))] = r.get("Value")

    claude = {}
    if "Claude_Summary" in names:
        for r in _rows(wb["Claude_Summary"]):
            claude[str(r.get("Metric"))] = r.get("Claude claims")

    trades = []
    for r in _rows(wb["Trades_BS"]):
        date = str(r.get("Date") or "")[:10]
        tin = str(r.get("In") or "")
        tout = str(r.get("Out") or "")
        e_spot, x_spot = _num(r.get("Entry Spot")), _num(r.get("Exit Spot"))
        trades.append({
            "side": str(r.get("Side") or "").lower(),
            "opt_type": r.get("Opt"), "strike": r.get("Strike"),
            "entry_dt": (date + " " + tin).strip(),
            "exit_dt": (date + " " + tout).strip(),
            "entry_prem": _num(r.get("Entry Prem")), "exit_prem": _num(r.get("Exit Prem")),
            "points": _num(r.get("Index Pts")), "qty": _num(r.get("Qty")),
            # spot levels (chart markers + Entry/Exit spot cols); alias to entry/exit
            "entry_spot": e_spot, "exit_spot": x_spot, "entry": e_spot, "exit": x_spot,
            "gross": _num(r.get("Gross Rs")), "fee": _num(r.get("Fee Rs")),
            "pnl": _num(r.get("Net Rs")), "bars": int(_num(r.get("Bars")) or 0),
            "reason": r.get("Exit Reason") or "",
        })
    return {"meta": meta, "trades": trades, "claude_summary": claude}


def recompute(trades):
    """Independent metric recompute from raw trades, via engine.metrics."""
    tr = [t for t in trades if t.get("pnl") is not None]
    tr = sorted(tr, key=lambda t: t["exit_dt"])
    if not tr:
        return {}
    dts = pd.to_datetime([t["exit_dt"] for t in tr], errors="coerce")
    eq = engine.START_CAP + pd.Series([t["pnl"] for t in tr]).cumsum()
    equity = pd.DataFrame({"Datetime": dts, "equity": eq.values})
    res = {"trades": [{"side": t["side"], "pnl": t["pnl"], "bars": t.get("bars", 0),
                       "exit": t.get("exit_prem") or 0, "qty": t.get("qty") or 0} for t in tr],
           "equity": equity, "final": float(eq.iloc[-1]),
           "variant": "xlsx", "mode": "intraday", "params": {}}
    m, _ = engine.metrics(res)
    m["fees"] = round(sum(t.get("fee", 0) or 0 for t in trades if t.get("fee")), 0)
    return m


def crosscheck(claude_summary, recomputed):
    """Rows: metric, claude, lab, delta, status, kind. Plus the CAGR-vs-net% note.

    kind='hard'  = reproducible from the trade list alone -> must match, else MISMATCH.
    kind='soft'  = path/sampling based (Sharpe/Sortino need the full calendar-day
                   equity, which is NOT in the trade list) -> shown for information,
                   never a hard mismatch (so a red flag always means a REAL problem).
    """
    pairs = [
        ("Trades", "trades", "hard"),
        ("Net %  (fixed-1-lot, NOT CAGR)", "net_pct", "hard"),
        ("Net Rs", "net_abs", "hard"),
        ("Profit Factor", "profit_factor", "hard"),
        ("Win rate %", "win_rate", "hard"),
        ("Max DD %", "maxdd", "hard"),
        ("Expectancy Rs", "expectancy", "hard"),
        ("Sharpe", "sharpe", "soft"),
    ]
    out = []
    for label, rkey, kind in pairs:
        cval = _num(claude_summary.get(label))
        rval = _num(recomputed.get(rkey))
        delta = None
        if cval is None or rval is None:
            status = "n/a"
        elif kind == "soft":
            delta = round(rval - cval, 3)
            status = "info (path-based)"
        else:
            delta = round(rval - cval, 3)
            tol = _TOL.get(rkey, 0.5)
            status = "match" if abs(delta) <= tol else "MISMATCH"
        out.append({"metric": label, "claude": cval, "lab": rval,
                    "delta": delta, "status": status, "kind": kind})
    # definitional cross-check: the "net %" vs a real CAGR (annual_return)
    out.append({"metric": "-> real CAGR (annual_return)", "claude": None,
                "lab": round(_num(recomputed.get("annual_return")) or 0, 2),
                "delta": None, "kind": "note",
                "status": "note: 'net %' is fixed-1-lot, NOT this CAGR (TRAP-127)"})
    return out


# Claude_Summary label -> metric key, for overlaying the run's CLAIMED headline
# numbers onto the display (so the uploaded dashboard shows the strategy's actual
# scored metrics, e.g. calendar-day Sharpe, not the trade-day recompute — the
# recompute stays in the cross-check panel where independence matters).
_CLAIMED = {
    "Net %  (fixed-1-lot, NOT CAGR)": "net_pct", "Net Rs": "net_abs",
    "Sharpe": "sharpe", "Sortino": "sortino", "Profit Factor": "profit_factor",
    "Win rate %": "win_rate", "Max DD %": "maxdd", "Expectancy Rs": "expectancy",
    "Trades": "trades",
}


def _downsample(arr, n=120):
    if len(arr) <= n:
        return [round(float(x), 1) for x in arr]
    idx = np.linspace(0, len(arr) - 1, n).astype(int)
    return [round(float(arr[i]), 1) for i in idx]


def monte_carlo(trades, meta_in, n_paths=1000, seed=7):
    """Bootstrap the per-trade P&L sequence 1000x -> net%/maxDD/Sharpe spread + fan.
    Genuinely trade-derived (unlike significance/opt-sweep which need the original
    signal/param search). Seeded so re-uploading the same file is reproducible."""
    pnl = np.array([float(t["pnl"]) for t in trades if t.get("pnl") is not None])
    if len(pnl) < 20:
        return None
    N = len(pnl)
    rng = np.random.default_rng(seed)
    try:
        win = str(meta_in.get("Window") or "").split(" -> ")
        years = max((pd.to_datetime(win[1]) - pd.to_datetime(win[0])).days / 365.25, 0.1)
    except Exception:
        years = max(N / 250.0, 0.1)
    tpy = N / years                                   # trades/year for annualisation
    cap = engine.START_CAP

    def _net(eq): return (eq[-1] - cap) / cap * 100.0
    def _dd(eq):
        peak = np.maximum.accumulate(eq)
        return float(((eq / peak - 1) * 100).min())
    def _sh(s):
        sd = s.std()
        return float(s.mean() / sd * np.sqrt(tpy)) if sd else 0.0

    nets, dds, shs, paths = [], [], [], []
    for i in range(n_paths):
        s = pnl[rng.integers(0, N, N)]
        eq = cap + np.cumsum(s)
        nets.append(_net(eq)); dds.append(_dd(eq)); shs.append(_sh(s))
        if i < 60:
            paths.append(_downsample(eq))
    orig_eq = cap + np.cumsum(pnl)

    def tbl(orig, arr):
        a = np.array(arr)
        return [round(float(orig), 2), round(float(np.percentile(a, 5)), 2),
                round(float(np.percentile(a, 50)), 2), round(float(np.percentile(a, 95)), 2)]

    return {"table": {"net": tbl(_net(orig_eq), nets), "maxdd": tbl(_dd(orig_eq), dds),
                      "sharpe": tbl(_sh(pnl), shs)},
            "paths": paths, "orig_path": _downsample(orig_eq)}


def to_results_payload(parsed, recomputed):
    """Build a minimal window.RESULTS-shaped dict the lab dashboard can render.
    Metrics = recompute (full set) with the run's CLAIMED headline values overlaid
    so the visual matches the original run; the cross-check panel keeps recompute."""
    metrics = dict(recomputed)
    for label, key in _CLAIMED.items():
        v = _num(parsed.get("claude_summary", {}).get(label))
        if v is not None:
            metrics[key] = v
    trades = [t for t in parsed["trades"] if t.get("pnl") is not None]
    trades = sorted(trades, key=lambda t: t["exit_dt"])
    # equity + labels
    eq = [engine.START_CAP]
    labels = [""]
    run = engine.START_CAP
    for t in trades:
        run += t["pnl"]; eq.append(round(run, 1)); labels.append(t["exit_dt"][:10])
    # monthly returns %
    monthly = {}
    for t in trades:
        y = t["exit_dt"][:4]; mo = int(t["exit_dt"][5:7]) if t["exit_dt"][5:7].isdigit() else 0
        if not mo:
            continue
        monthly.setdefault(y, {})
        monthly[y][mo] = round(monthly[y].get(mo, 0) + t["pnl"] / engine.START_CAP * 100, 3)
    meta_in = parsed["meta"]
    win = str(meta_in.get("Window") or "").split(" -> ")
    meta = {"window": win if len(win) == 2 else [labels[1] if len(labels) > 1 else "", labels[-1]],
            "days": _num(meta_in.get("Trading days")) or len(trades),
            "start_cap": engine.START_CAP,
            "design": meta_in.get("Strategy") or "Uploaded Excel",
            "tf": meta_in.get("Timeframe") or "", "instrument": meta_in.get("Instrument") or "",
            "lot_size": _num(meta_in.get("Lot size")) or 65, "lots": _num(meta_in.get("Lots")) or 1,
            "candles": [], "passes": ["bs"], "periods": ["full"]}
    combo = {"dna": {}, "metrics": metrics, "all_trades": trades,
             "trades": trades[-10:], "equity": eq, "benchmark": eq, "labels": labels,
             "underwater": [], "worst_periods": [], "monthly": monthly,
             "mc": monte_carlo(trades, parsed.get("meta", {}))}
    # significance + opt_table deliberately absent — they need the original run's
    # signal/param search, not reproducible from the trade list; the dashboard now
    # hides those two panels when absent.
    return {"meta": meta, "combos": {"bs|full": combo}}


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "runs/mid_orb_nifty/mid_orb_nifty.xlsx"
    parsed = parse(p)
    rc = recompute(parsed["trades"])
    print("parsed {} trades".format(len(parsed["trades"])))
    print("recompute: net_abs={} net_pct={:.2f} sharpe={:.2f} pf={:.2f} win={:.1f} dd={:.2f}".format(
        rc.get("net_abs"), rc.get("net_pct"), rc.get("sharpe"), rc.get("profit_factor"),
        rc.get("win_rate"), rc.get("maxdd")))
    print("CROSS-CHECK:")
    for row in crosscheck(parsed["claude_summary"], rc):
        print("  {:36} claude={} lab={} -> {}".format(
            row["metric"], row["claude"], row["lab"], row["status"]))
